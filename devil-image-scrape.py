"""Bulk-resolve Devil Mountain Nursery product image URLs for the KLR plant catalog.

One HTTP GET per product page; og:image is in <head>, so the parse is cheap and
no JS execution is needed. The CDN path fragment (pid/iid/file) is stored rather
than a single rendered size because BigCommerce serves any WxH from the same
fragment -- storing the fragment means a resolution change never requires a recrawl.
"""

from __future__ import annotations

import argparse
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Optional

import openpyxl
import requests

OG_IMAGE = re.compile(
    r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']', re.I
)
# .../s-<store>/products/<pid>/images/<iid>/<file>.<w>.<h>.jpg?c=1
OG_PATH = re.compile(
    r"/(s-[a-z0-9]+)/products/(\d+)/images/(\d+)/(.+?)\.\d+\.\d+\.(jpe?g|png|webp)", re.I
)
STENCIL = "https://cdn11.bigcommerce.com/{store}/images/stencil/{size}/products/{pid}/{iid}/{file}.{ext}?c=1"

HEADERS = {"User-Agent": "KLRBuild-CatalogBot/1.0 (+nicolasdonatelli@gmail.com)"}


def resolve(url: str, session: requests.Session, size: str) -> tuple[Optional[str], Optional[str]]:
    """Return (canonical_image_url, cdn_path_key) or (None, None).

    Streams the response and stops reading after </head> to avoid
    downloading the full page body (~80 % bandwidth savings).
    """
    try:
        resp = session.get(url, headers=HEADERS, timeout=20, stream=True)
        resp.raise_for_status()
    except requests.RequestException:
        return None, None

    # Read in small chunks and stop once we have the full <head>
    buf: list[str] = []
    head_html = ""
    try:
        for chunk in resp.iter_content(chunk_size=4096, decode_unicode=True):
            if chunk:
                buf.append(chunk)
                joined = "".join(buf)
                if "</head>" in joined.lower():
                    head_html = joined
                    break
        else:
            # Never found </head>; use whatever we collected
            head_html = "".join(buf)
    finally:
        resp.close()

    match = OG_IMAGE.search(head_html)
    if not match:
        return None, None

    parts = OG_PATH.search(match.group(1))
    if not parts:
        return match.group(1), None  # non-standard asset; keep raw

    store, pid, iid, file, ext = parts.groups()
    key = f"{pid}/{iid}/{file}.{ext}"
    return STENCIL.format(store=store, size=size, pid=pid, iid=iid, file=file, ext=ext), key


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path)
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--sheet", default="All Plants")
    ap.add_argument("--size", default="1280x1280")
    ap.add_argument("--delay", type=float, default=1.0,
                    help="Per-thread delay between requests (default: 1s)")
    ap.add_argument("--workers", type=int, default=3,
                    help="Number of concurrent workers (default: 3)")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--overwrite", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook)
    ws = wb[args.sheet]

    # BH=60 Source URL, BI=61 Image URL, BJ=62 GBIF, BK=63 Image Source, BL=64 CDN Path Key
    ws.cell(row=1, column=63, value="Image Source")
    ws.cell(row=1, column=64, value="CDN Path Key")

    # Collect rows that need resolving
    work: list[tuple[int, str]] = []  # (row_number, url)
    for row in range(2, ws.max_row + 1):
        src = ws.cell(row=row, column=60).value
        if not isinstance(src, str) or "devilmountainnursery.com" not in src:
            continue
        if ws.cell(row=row, column=61).value and not args.overwrite:
            continue
        work.append((row, src.strip()))
        if args.limit and len(work) >= args.limit:
            break

    total = len(work)
    if total == 0:
        print("nothing to resolve")
        wb.save(args.out)
        return

    print(f"resolving {total} rows with {args.workers} workers, {args.delay}s delay …")

    wb_lock = threading.Lock()
    done = 0
    t0 = time.monotonic()

    def _worker(row: int, url: str) -> tuple[int, Optional[str], Optional[str]]:
        session = requests.Session()
        img, key = resolve(url, session, args.size)
        time.sleep(args.delay)
        return row, img, key

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(_worker, r, u): r for r, u in work}
        for future in as_completed(futures):
            row, img, key = future.result()
            with wb_lock:
                if img:
                    ws.cell(row=row, column=61, value=img)
                    ws.cell(row=row, column=63, value="DMN")
                    ws.cell(row=row, column=64, value=key)
                else:
                    ws.cell(row=row, column=61, value="no image")
                    gbif = ws.cell(row=row, column=62).value
                    ws.cell(row=row, column=63, value="GBIF" if gbif else "none")

                done += 1
                if done % 25 == 0:
                    wb.save(args.out)  # checkpoint
                    elapsed = time.monotonic() - t0
                    rate = done / elapsed
                    eta = (total - done) / rate if rate > 0 else 0
                    print(f"{done}/{total} resolved  ({rate:.1f}/s, ~{eta:.0f}s remaining)",
                          flush=True)

    wb.save(args.out)
    elapsed = time.monotonic() - t0
    print(f"done: {done} rows written -> {args.out}  ({elapsed:.1f}s)")


if __name__ == "__main__":
    main()